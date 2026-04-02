import os
import pandas as pd
from openpyxl.styles import PatternFill


def apply_color_coding(ws, df):
    fills = {
        "concordant": PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE"),
        "minor_discordant": PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C"),
        "major_discordant": PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"),
    }

    if "Concordance_Class" not in df.columns:
        print("Concordance_Class column not found.")
        return

    excel_col_idx = df.columns.get_loc("Concordance_Class") + 1

    for row in range(2, len(df) + 2):
        value = ws.cell(row=row, column=excel_col_idx).value
        if value in fills:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fills[value]


def export_results(classified_df, results, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        classified_df.to_excel(writer, sheet_name="Classified_Cases", index=False)
        results["concordance_table"].to_excel(writer, sheet_name="Concordance_Summary", index=False)
        results["subtype_table"].to_excel(writer, sheet_name="Subtype_Summary", index=False)
        results["direction_table"].to_excel(writer, sheet_name="Direction_Summary", index=False)
        results["severity_table"].to_excel(writer, sheet_name="Severity_Summary", index=False)
        results["confusion_matrix"].to_excel(writer, sheet_name="Confusion_Matrix")
        results["hsil_pv_table"].to_excel(writer, sheet_name="HSIL_PV_Plus", index=False)

        ws = writer.sheets["Classified_Cases"]
        apply_color_coding(ws, classified_df)

    print(f"Exported with color coding: {output_path}")
